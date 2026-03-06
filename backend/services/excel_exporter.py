"""Generación de archivos Excel (.xlsx) con openpyxl."""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from config import EXPORTS_DIR
from services.hmm_trainer import (
    get_emission_probs, get_transition_probs,
    get_emission_table, get_transition_table,
)
from services.corpus_parser import get_corpus_data
from services.eagles_tags import describe_tag

logger = logging.getLogger(__name__)

# Estilos
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
HIGHLIGHT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def _style_header(ws, row, max_col):
    """Aplica estilo de encabezado a una fila."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _auto_width(ws, min_width=8, max_width=30):
    """Ajusta el ancho de columnas automáticamente."""
    for col in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[col_letter].width = max_len


def generate_emission_excel(top_n: int = 30) -> str:
    """Genera Excel con tabla de probabilidades de emisión.

    Returns:
        Ruta del archivo generado.
    """
    emission_table = get_emission_table(top_n)
    if not emission_table:
        raise RuntimeError("No hay datos de emisión disponibles.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Probabilidades de Emisión"

    # Título
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "Tabla de Probabilidades de Emisión P(palabra|etiqueta)"
    title_cell.font = Font(bold=True, size=14, color="2F5496")
    title_cell.alignment = CENTER

    ws.merge_cells("A2:E2")
    ws["A2"].value = "HMM Bigrama — Wikicorpus Español (EAGLES/FreeLing)"
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = CENTER

    # Encabezados
    headers = ["Etiqueta", "Descripción", "Total Tag", "Palabra", "Conteo", "P(palabra|etiqueta)"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    _style_header(ws, 4, len(headers))

    # Datos
    row = 5
    for entry in emission_table:
        tag = entry["tag"]
        tag_desc = describe_tag(tag)["description"]
        tag_count = entry["tag_count"]

        for i, word_data in enumerate(entry["top_words"][:10]):
            ws.cell(row=row, column=1, value=tag if i == 0 else "")
            ws.cell(row=row, column=2, value=tag_desc if i == 0 else "")
            ws.cell(row=row, column=3, value=tag_count if i == 0 else "")
            ws.cell(row=row, column=4, value=word_data["word"])
            ws.cell(row=row, column=5, value=word_data["count"])
            ws.cell(row=row, column=6, value=word_data["probability"])
            ws.cell(row=row, column=6).number_format = "0.00000000"

            if i == 0:
                for c in range(1, 4):
                    ws.cell(row=row, column=c).fill = SUBHEADER_FILL
                    ws.cell(row=row, column=c).font = Font(bold=True)

            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).border = BORDER
                ws.cell(row=row, column=c).alignment = CENTER if c != 4 else LEFT

            row += 1

    _auto_width(ws)

    filepath = EXPORTS_DIR / "tabla_emision.xlsx"
    wb.save(str(filepath))
    logger.info(f"Excel de emisión generado: {filepath}")
    return str(filepath)


def generate_transition_excel() -> str:
    """Genera Excel con tabla de probabilidades de transición.

    Returns:
        Ruta del archivo generado.
    """
    transition_table = get_transition_table()
    if not transition_table:
        raise RuntimeError("No hay datos de transición disponibles.")

    data = get_corpus_data()
    tag_counts = data["tag_counts"] if data else {}
    sorted_tags = sorted(tag_counts.keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Probabilidades de Transición"

    # Título
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = "Tabla de Probabilidades de Transición P(etiqueta_i|etiqueta_{i-1})"
    title_cell.font = Font(bold=True, size=14, color="2F5496")
    title_cell.alignment = CENTER

    ws.merge_cells("A2:E2")
    ws["A2"].value = "HMM Bigrama — Wikicorpus Español (EAGLES/FreeLing)"
    ws["A2"].font = Font(italic=True, size=10)
    ws["A2"].alignment = CENTER

    # Formato de lista
    headers = ["Etiqueta Previa", "Etiqueta Siguiente", "Conteo", "P(tag_i|tag_{i-1})"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    _style_header(ws, 4, len(headers))

    row = 5
    for entry in transition_table[:2000]:  # Limitar a 2000 filas
        ws.cell(row=row, column=1, value=entry["tag_prev"])
        ws.cell(row=row, column=2, value=entry["tag_next"])
        ws.cell(row=row, column=3, value=entry["count"])
        ws.cell(row=row, column=4, value=entry["probability"])
        ws.cell(row=row, column=4).number_format = "0.00000000"

        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).alignment = CENTER
        row += 1

    # Hoja 2: Matriz de transición (top 30 tags)
    ws2 = wb.create_sheet("Matriz de Transición")
    top_tags = ["<START>"] + sorted(tag_counts, key=tag_counts.get, reverse=True)[:30] + ["<END>"]

    transition_probs = get_transition_probs()
    if transition_probs:
        ws2.cell(row=1, column=1, value="P(col|fila)")
        ws2.cell(row=1, column=1).font = Font(bold=True)
        for j, tag in enumerate(top_tags):
            ws2.cell(row=1, column=j + 2, value=tag)
            ws2.cell(row=1, column=j + 2).font = HEADER_FONT
            ws2.cell(row=1, column=j + 2).fill = HEADER_FILL
            ws2.cell(row=1, column=j + 2).alignment = CENTER

        for i, prev_tag in enumerate(top_tags):
            ws2.cell(row=i + 2, column=1, value=prev_tag)
            ws2.cell(row=i + 2, column=1).font = Font(bold=True)
            ws2.cell(row=i + 2, column=1).fill = SUBHEADER_FILL

            for j, next_tag in enumerate(top_tags):
                prob = transition_probs.get((prev_tag, next_tag), 0)
                cell = ws2.cell(row=i + 2, column=j + 2, value=prob if prob > 0 else "")
                if prob > 0:
                    cell.number_format = "0.0000"
                cell.border = BORDER
                cell.alignment = CENTER
                if prob > 0.1:
                    cell.fill = HIGHLIGHT_FILL

        _auto_width(ws2, min_width=6, max_width=12)

    _auto_width(ws)

    filepath = EXPORTS_DIR / "tabla_transicion.xlsx"
    wb.save(str(filepath))
    logger.info(f"Excel de transición generado: {filepath}")
    return str(filepath)


def generate_viterbi_excel(viterbi_result: dict, filename: str = "viterbi_resultado.xlsx") -> str:
    """Genera Excel con la matriz de Viterbi de un etiquetado.

    Args:
        viterbi_result: Resultado del algoritmo de Viterbi.
        filename: Nombre del archivo de salida.

    Returns:
        Ruta del archivo generado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultado Viterbi"

    tokens = viterbi_result["tokens"]
    tags = viterbi_result["tags"]
    descriptions = viterbi_result["descriptions"]

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(tokens) + 1)
    ws["A1"].value = f"Algoritmo de Viterbi: \"{viterbi_result['sentence']}\""
    ws["A1"].font = Font(bold=True, size=14, color="2F5496")

    # Fila de tokens
    ws.cell(row=3, column=1, value="Token")
    for j, token in enumerate(tokens):
        ws.cell(row=3, column=j + 2, value=token)
    _style_header(ws, 3, len(tokens) + 1)

    # Fila de etiquetas asignadas
    ws.cell(row=4, column=1, value="Etiqueta")
    ws.cell(row=4, column=1).font = Font(bold=True)
    ws.cell(row=4, column=1).fill = SUBHEADER_FILL
    for j, tag in enumerate(tags):
        cell = ws.cell(row=4, column=j + 2, value=tag)
        cell.fill = HIGHLIGHT_FILL
        cell.font = Font(bold=True)
        cell.border = BORDER
        cell.alignment = CENTER

    # Fila de descripciones
    ws.cell(row=5, column=1, value="Descripción")
    ws.cell(row=5, column=1).font = Font(bold=True)
    ws.cell(row=5, column=1).fill = SUBHEADER_FILL
    for j, desc in enumerate(descriptions):
        cell = ws.cell(row=5, column=j + 2, value=desc)
        cell.font = Font(italic=True, size=9)
        cell.border = BORDER
        cell.alignment = LEFT

    # Matriz de Viterbi
    ws.cell(row=7, column=1, value="Matriz de Viterbi (probabilidades)")
    ws.cell(row=7, column=1).font = Font(bold=True, size=12, color="2F5496")

    viterbi_matrix = viterbi_result["viterbi_matrix"]
    if viterbi_matrix:
        # Recopilar todas las etiquetas presentes en la matriz
        all_matrix_tags = set()
        for row_data in viterbi_matrix:
            for key in row_data:
                if key != "token":
                    all_matrix_tags.add(key)
        sorted_matrix_tags = sorted(all_matrix_tags)

        # Encabezado
        ws.cell(row=8, column=1, value="Etiqueta \\ Token")
        for j, token in enumerate(tokens):
            ws.cell(row=8, column=j + 2, value=token)
        _style_header(ws, 8, len(tokens) + 1)

        # Filas por etiqueta
        for i, tag in enumerate(sorted_matrix_tags):
            row_num = 9 + i
            ws.cell(row=row_num, column=1, value=tag)
            ws.cell(row=row_num, column=1).font = Font(bold=True)
            ws.cell(row=row_num, column=1).fill = SUBHEADER_FILL
            ws.cell(row=row_num, column=1).border = BORDER

            for j, row_data in enumerate(viterbi_matrix):
                val = row_data.get(tag, "")
                cell = ws.cell(row=row_num, column=j + 2)
                if val != "":
                    cell.value = val
                    cell.number_format = "0.00E+00"
                cell.border = BORDER
                cell.alignment = CENTER

                # Resaltar la ruta óptima
                if j < len(tags) and tags[j] == tag and val != "":
                    cell.fill = HIGHLIGHT_FILL

    _auto_width(ws)

    filepath = EXPORTS_DIR / filename
    wb.save(str(filepath))
    logger.info(f"Excel de Viterbi generado: {filepath}")
    return str(filepath)


def generate_comparison_excel(result_1: dict, result_2: dict) -> str:
    """Genera Excel con la comparación de dos etiquetados de Viterbi.

    Args:
        result_1: Resultado de viterbi() para la oración 1.
        result_2: Resultado de viterbi() para la oración 2.

    Returns:
        Ruta del archivo generado.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparación"

    # Título
    ws.merge_cells("A1:G1")
    ws["A1"].value = "Comparación de Etiquetados — Oraciones Obligatorias"
    ws["A1"].font = Font(bold=True, size=14, color="2F5496")
    ws["A1"].alignment = CENTER

    # ── Oración 1 ──
    ws.merge_cells("A3:G3")
    ws["A3"].value = f"Oración 1: \"{result_1['sentence']}\""
    ws["A3"].font = Font(bold=True, size=11)

    headers = ["Posición", "Token", "Etiqueta", "Descripción"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    _style_header(ws, 4, len(headers))

    row = 5
    for j, (tok, tag, desc) in enumerate(zip(
        result_1["tokens"], result_1["tags"], result_1["descriptions"]
    )):
        ws.cell(row=row, column=1, value=j + 1)
        ws.cell(row=row, column=2, value=tok)
        ws.cell(row=row, column=3, value=tag)
        ws.cell(row=row, column=4, value=desc)
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).alignment = CENTER if c <= 3 else LEFT
        row += 1

    # ── Oración 2 ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1).value = f"Oración 2: \"{result_2['sentence']}\""
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    row += 1

    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header(ws, row, len(headers))
    row += 1

    for j, (tok, tag, desc) in enumerate(zip(
        result_2["tokens"], result_2["tags"], result_2["descriptions"]
    )):
        ws.cell(row=row, column=1, value=j + 1)
        ws.cell(row=row, column=2, value=tok)
        ws.cell(row=row, column=3, value=tag)
        ws.cell(row=row, column=4, value=desc)
        for c in range(1, 5):
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).alignment = CENTER if c <= 3 else LEFT
        row += 1

    # ── Tabla comparativa ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    ws.cell(row=row, column=1).value = "Comparación por Palabra Compartida"
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="2F5496")
    row += 1

    comp_headers = ["Palabra", "Etiqueta Or.1", "Descripción Or.1",
                    "Etiqueta Or.2", "Descripción Or.2", "¿Coinciden?"]
    for col, h in enumerate(comp_headers, 1):
        ws.cell(row=row, column=col, value=h)
    _style_header(ws, row, len(comp_headers))
    row += 1

    # Buscar palabras comunes
    map_1 = {tok.lower(): (tag, desc) for tok, tag, desc in zip(
        result_1["tokens"], result_1["tags"], result_1["descriptions"]
    )}
    map_2 = {tok.lower(): (tag, desc) for tok, tag, desc in zip(
        result_2["tokens"], result_2["tags"], result_2["descriptions"]
    )}
    palabras_comunes = [w for w in map_1 if w in map_2]

    match_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    diff_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

    for palabra in palabras_comunes:
        tag1, desc1 = map_1[palabra]
        tag2, desc2 = map_2[palabra]
        coincide = tag1 == tag2
        ws.cell(row=row, column=1, value=palabra)
        ws.cell(row=row, column=2, value=tag1)
        ws.cell(row=row, column=3, value=desc1)
        ws.cell(row=row, column=4, value=tag2)
        ws.cell(row=row, column=5, value=desc2)
        ws.cell(row=row, column=6, value="Sí" if coincide else "No")

        fill = match_fill if coincide else diff_fill
        for c in range(1, 7):
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).alignment = CENTER if c in (2, 4, 6) else LEFT
            ws.cell(row=row, column=c).fill = fill
        row += 1

    # ── Resumen de probabilidades ──
    row += 1
    ws.cell(row=row, column=1, value="Log-probabilidad Oración 1:")
    ws.cell(row=row, column=1).font = Font(bold=True)
    import math
    lp1 = math.log(result_1["best_path_prob"]) if result_1["best_path_prob"] > 0 else 0
    lp2 = math.log(result_2["best_path_prob"]) if result_2["best_path_prob"] > 0 else 0
    ws.cell(row=row, column=2, value=lp1)
    ws.cell(row=row, column=2).number_format = "0.0000"
    row += 1
    ws.cell(row=row, column=1, value="Log-probabilidad Oración 2:")
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2, value=lp2)
    ws.cell(row=row, column=2).number_format = "0.0000"
    row += 1
    mas_probable = "Oración 2" if lp2 > lp1 else "Oración 1"
    ws.cell(row=row, column=1, value=f"Más probable: {mas_probable}")
    ws.cell(row=row, column=1).font = Font(bold=True, color="2F5496")

    _auto_width(ws)

    filepath = EXPORTS_DIR / "comparacion_etiquetados.xlsx"
    wb.save(str(filepath))
    logger.info(f"Excel de comparación generado: {filepath}")
    return str(filepath)
